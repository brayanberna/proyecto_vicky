from django.shortcuts import render
from django.shortcuts import redirect

# Libreria Excel
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
#from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Libreria Fecha
from datetime import datetime

import numpy as np
# Create your views here.
def export(request):
  myDate=datetime.now()
  fecha = myDate.strftime("%d-%m-%Y")
  cant_masculino = 0
  cant_femenino = 0
  pauta_breve_hombres = 0
  pauta_breve_mujeres = 0
  # SECCIÓN A.1: APLICACIÓN Y RESULTADOS DE PAUTA BREVE
  
  if request.method == 'POST':
    #print(request.FILES['excel1'])
    #print(request.FILES['excel2'])
    # Excel con datos
    workbook_name = request.FILES['excel1']
    workbook_name2 = request.FILES['excel2']
    wb = load_workbook(workbook_name)
    wb2 = load_workbook(workbook_name2)
    ws = wb[wb.sheetnames[-3]]
    ws2 = wb2[wb2.sheetnames[0]]
    print("Hoja 1:",wb.sheetnames[-3])
    print("Hoja 2:",wb2.sheetnames[0])

    for rows in ws['A1':get_column_letter(ws.max_column) + '1']:
      for cell in rows:
        if cell.value == "SEXO":
          #print("SEXO:",cell.coordinate)
          letra_columna_sexo = cell.column_letter
        if cell.value == "RANGO ETAREO":
          #print("RANGO ETARIO:",cell.coordinate)
          letra_columna_rango_etario = cell.column_letter
        if cell.value == "EDAD":
          #print("EDAD:",cell.coordinate)
          letra_columna_edad = cell.column_letter
        if cell.value == "PAUTA DSM":
          #print("PAUTA DSM:",cell.coordinate)
          letra_columna_pauta_dsm = cell.column_letter

    cant_filas = ws.max_row

    for rows in ws[letra_columna_sexo + '2':letra_columna_sexo + str(cant_filas)]:
      for cell in rows:
        if cell.value == "MASCULINO":
          cant_masculino = cant_masculino + 1
        if cell.value == "FEMENINO":
          cant_femenino = cant_femenino + 1

    # SECCIÓN A.1: APLICACIÓN Y RESULTADOS DE PAUTA BREVE
    menor_un_mes_f = 0
    un_mes_f = 0
    dos_meses_f = 0
    tres_meses_f = 0
    cuatro_meses_f = 0
    cinco_meses_f = 0
    seis_meses_f = 0
    siete_once_meses_f = 0
    doce_diecisiete_meses_f = 0
    dieciocho_veintitres_meses_f = 0
    menor_un_mes_m = 0
    un_mes_m = 0
    dos_meses_m = 0
    tres_meses_m = 0
    cuatro_meses_m = 0
    cinco_meses_m = 0
    seis_meses_m = 0
    siete_once_meses_m = 0
    doce_diecisiete_meses_m = 0
    dieciocho_veintitres_meses_m = 0
    for rows in ws[letra_columna_pauta_dsm + '2':letra_columna_pauta_dsm + str(cant_filas)]:
      for cell in rows:
        if cell.value != None:
          if cell.value.upper() == "PAUTA BREVE":
            columna_rango_etario = ws[letra_columna_rango_etario + str(cell.row)].value.upper()
            column_edad = ws[letra_columna_edad + str(cell.row)].value.upper().split()
            if ws[letra_columna_sexo + str(cell.row)].value == 'MASCULINO':
              if columna_rango_etario == "MENOR DE 1 MES" or column_edad[1] == "DIAS" or column_edad[1] == "DIA" or column_edad[1] == "D":
                menor_un_mes_m = menor_un_mes_m + 1
                pauta_breve_hombres = pauta_breve_hombres + 1
              if columna_rango_etario == "1 MES":
                un_mes_m = un_mes_m + 1
                pauta_breve_hombres = pauta_breve_hombres + 1
              if columna_rango_etario == "2 MESES":
                dos_meses_m = dos_meses_m + 1
                pauta_breve_hombres = pauta_breve_hombres + 1
              if columna_rango_etario == "3 MESES":
                tres_meses_m = tres_meses_m + 1
                pauta_breve_hombres = pauta_breve_hombres + 1
              if columna_rango_etario == "4 MESES":
                cuatro_meses_m = cuatro_meses_m + 1
                pauta_breve_hombres = pauta_breve_hombres + 1
              if columna_rango_etario == "5 MESES":
                cinco_meses_m = cinco_meses_m + 1
                pauta_breve_hombres = pauta_breve_hombres + 1
              if columna_rango_etario == "6 MESES" or column_edad[0] + ' ' + column_edad[1] == "6 MESES":
                seis_meses_m = seis_meses_m + 1
                pauta_breve_hombres = pauta_breve_hombres + 1
              if columna_rango_etario == "7 A 11 MESES" or column_edad[0] + ' ' + column_edad[1] == "7 MESES" or column_edad[0] + ' ' + column_edad[1] == "8 MESES" or column_edad[0] + ' ' + column_edad[1] == "9 MESES" or column_edad[0] + ' ' + column_edad[1] == "10 MESES" or column_edad[0] + ' ' + column_edad[1] == "11 MESES":
                siete_once_meses_m = siete_once_meses_m + 1
                pauta_breve_hombres = pauta_breve_hombres + 1
              if columna_rango_etario == "12 A 17 MESES" or column_edad[0] + ' ' + column_edad[1] == "12 MESES" or column_edad[0] + ' ' + column_edad[1] == "13 MESES" or column_edad[0] + ' ' + column_edad[1] == "14 MESES" or column_edad[0] + ' ' + column_edad[1] == "15 MESES" or column_edad[0] + ' ' + column_edad[1] == "16 MESES" or column_edad[0] + ' ' + column_edad[1] == "17 MESES":
                doce_diecisiete_meses_m = doce_diecisiete_meses_m + 1
                pauta_breve_hombres = pauta_breve_hombres + 1
              if columna_rango_etario == "18 A 23 MESES" or columna_rango_etario == " 2 AÑOS" or column_edad[0] + ' ' + column_edad[1] == "18 MESES" or column_edad[0] + ' ' + column_edad[1] == "19 MESES" or column_edad[0] + ' ' + column_edad[1] == "20 MESES" or column_edad[0] + ' ' + column_edad[1] == "21 MESES" or column_edad[0] + ' ' + column_edad[1] == "22 MESES" or column_edad[0] + ' ' + column_edad[1] == "23 MESES" or column_edad[0] + ' ' + column_edad[1] == "24 MESES":
                dieciocho_veintitres_meses_m = dieciocho_veintitres_meses_m + 1
                pauta_breve_hombres = pauta_breve_hombres + 1
            if ws[letra_columna_sexo + str(cell.row)].value == 'FEMENINO':
              if columna_rango_etario == "MENOR DE 1 MES" or column_edad[1] == "DIAS" or column_edad[1] == "DIA" or column_edad[1] == "D":
                menor_un_mes_f = menor_un_mes_f + 1
                pauta_breve_mujeres = pauta_breve_mujeres + 1
              if columna_rango_etario == "1 MES":
                un_mes_f = un_mes_f + 1
                pauta_breve_mujeres = pauta_breve_mujeres + 1
              if columna_rango_etario == "2 MESES":
                dos_meses_f = dos_meses_f + 1
                pauta_breve_mujeres = pauta_breve_mujeres + 1
              if columna_rango_etario == "3 MESES":
                tres_meses_f = tres_meses_f + 1
                pauta_breve_mujeres = pauta_breve_mujeres + 1
              if columna_rango_etario == "4 MESES":
                cuatro_meses_f = cuatro_meses_f + 1
                pauta_breve_mujeres = pauta_breve_mujeres + 1
              if columna_rango_etario == "5 MESES":
                cinco_meses_f = cinco_meses_f + 1
                pauta_breve_mujeres = pauta_breve_mujeres + 1
              if columna_rango_etario == "6 MESES" or column_edad[0] + ' ' + column_edad[1] == "6 MESES":
                seis_meses_f = seis_meses_f + 1
                pauta_breve_mujeres = pauta_breve_mujeres + 1
              if columna_rango_etario == "7 A 11 MESES" or column_edad[0] + ' ' + column_edad[1] == "7 MESES" or column_edad[0] + ' ' + column_edad[1] == "8 MESES" or column_edad[0] + ' ' + column_edad[1] == "9 MESES" or column_edad[0] + ' ' + column_edad[1] == "10 MESES" or column_edad[0] + ' ' + column_edad[1] == "11 MESES":
                siete_once_meses_f = siete_once_meses_f + 1
                pauta_breve_mujeres = pauta_breve_mujeres + 1
              if columna_rango_etario == "12 A 17 MESES" or column_edad[0] + ' ' + column_edad[1] == "12 MESES" or column_edad[0] + ' ' + column_edad[1] == "13 MESES" or column_edad[0] + ' ' + column_edad[1] == "14 MESES" or column_edad[0] + ' ' + column_edad[1] == "15 MESES" or column_edad[0] + ' ' + column_edad[1] == "16 MESES" or column_edad[0] + ' ' + column_edad[1] == "17 MESES":
                doce_diecisiete_meses_f = doce_diecisiete_meses_f + 1
                pauta_breve_mujeres = pauta_breve_mujeres + 1
              if columna_rango_etario == "18 A 23 MESES" or columna_rango_etario == " 2 AÑOS" or column_edad[0] + ' ' + column_edad[1] == "18 MESES" or column_edad[0] + ' ' + column_edad[1] == "19 MESES" or column_edad[0] + ' ' + column_edad[1] == "20 MESES" or column_edad[0] + ' ' + column_edad[1] == "21 MESES" or column_edad[0] + ' ' + column_edad[1] == "22 MESES" or column_edad[0] + ' ' + column_edad[1] == "23 MESES" or column_edad[0] + ' ' + column_edad[1] == "24 MESES":
                dieciocho_veintitres_meses_f = dieciocho_veintitres_meses_f + 1
                pauta_breve_mujeres = pauta_breve_mujeres + 1

    #print("tablitas 1: ",ws.calculate_dimension()) # Dimensiones
    #print("tablitas 2: ",ws.max_row) # Cantidad Total de filas

    #Excel con calculos
    ws2['C34'] = pauta_breve_hombres + pauta_breve_mujeres
    ws2['D34'] = pauta_breve_hombres
    ws2['E34'] = pauta_breve_mujeres
    ws2['F34'] = menor_un_mes_m
    ws2['G34'] = menor_un_mes_f
    ws2['H34'] = un_mes_m
    ws2['I34'] = un_mes_f
    ws2['J34'] = dos_meses_m
    ws2['K34'] = dos_meses_f
    ws2['L34'] = tres_meses_m
    ws2['M34'] = tres_meses_f
    ws2['N34'] = cuatro_meses_m
    ws2['O34'] = cuatro_meses_f
    ws2['P34'] = cinco_meses_m
    ws2['Q34'] = cinco_meses_f
    ws2['R34'] = seis_meses_m
    ws2['S34'] = seis_meses_f
    ws2['T34'] = siete_once_meses_m
    ws2['U34'] = siete_once_meses_f
    ws2['V34'] = doce_diecisiete_meses_m
    ws2['W34'] = doce_diecisiete_meses_f
    ws2['X34'] = dieciocho_veintitres_meses_m
    ws2['Y34'] = dieciocho_veintitres_meses_f

    wb2.save('C:/Users/56975/Documents/FORMATO CONTROL NIÑO SANO.xlsx')
    return redirect('export')
  return render(request, 'export.html')
from django.shortcuts import render
from django.shortcuts import redirect

from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth import logout
from django.contrib.auth import authenticate

from django.contrib.auth.models import User

from .forms import RegisterForm, RegisterFormIndex

# Libreria Excel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# Libreria Fecha
from datetime import datetime
#import locale
# Idioma "es-ES" (código para el español de España)
#locale.setlocale(locale.LC_ALL, 'es-ES') 

# Create your views here.
myDate=datetime.now()

def index(request):
  form = RegisterFormIndex(request.POST or None)
  
  if request.method == 'POST' and form.is_valid():
    #formatedDate = myDate.strftime("%Y-%m-%d %H:%M:%S")
    localidad = form.cleaned_data.get('localidad')
    fecha = myDate.strftime("%d-%m-%Y")
    nombre = form.cleaned_data.get('nombre')
    rut = form.cleaned_data.get('rut')
    tipo_atencion = form.cleaned_data.get('tipo_atencion')
    nacionalidad = form.cleaned_data.get('nacionalidad')
    etnia = form.cleaned_data.get('etnia')
    sexo = form.cleaned_data.get('sexo')
    rango_etario = form.cleaned_data.get('rango_etario')
    edad = form.cleaned_data.get('edad')

    try:
      workbook_name = 'C:/Users/56975/Documents/proyecto_vicky.xlsx'
      wb = load_workbook(workbook_name)
      ws = wb.active
      print()
      if wb.sheetnames[-1] == (myDate.strftime("%B") + myDate.strftime("%Y")).upper():
        ws = wb[wb.sheetnames[-1]]
        #print("Fecha 1:",myDate.strftime("%Y"))
        #print("Fecha 2:",myDate.strftime("%B"))
        numero = int(ws.cell(row=ws.max_row, column=1).value) + 1
        datos = [numero, localidad, fecha, nombre, rut, tipo_atencion, nacionalidad, etnia, sexo, rango_etario, edad]
        ws.append(datos)
      
        for rows in ws['A2':max(ws.calculate_dimension()) + str(ws.max_row)]:
          for cell in rows:
            cell.font = Font(size=11)  #Color de texto
            bd = Side(style='thin', color="000000")
            cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)

        for rows in ws['A2': 'C' + str(ws.max_row)]:
          for cell in rows:
            cell.alignment = Alignment(horizontal="center")

        max_row=ws.max_row
        max_column=ws.max_column

        ws.auto_filter.ref='A1:' + max(ws.calculate_dimension()) + str(ws.max_row)
        wb.save('C:/Users/56975/Documents/proyecto_vicky.xlsx')
        return redirect('index')
        
      else:
        #print("Fecha 1:",myDate.strftime("%Y"))
        #print("Fecha 2:",myDate.strftime("%B"))
        ws = wb.create_sheet((myDate.strftime("%B") + myDate.strftime("%Y")).upper()) #Crea nuevas hojas de trabajo.
        columnas = ['N°', 'LOCALIDAD', 'FECHA', 'NOMBRE', 'RUT', 'TIPO ATENCION', 'NACIONALIDAD', 'ETNIA', 'SEXO', 'RANGO ETARIO', 'EDAD']
        ws.append(columnas)
        datos = ['1', localidad, fecha, nombre, rut, tipo_atencion, nacionalidad, etnia, sexo, rango_etario, edad]
        ws.append(datos)

        for rows in ws['A1':'K1']:
          for cell in rows:
            cell.font = Font(color="000000", size=16, bold=True)  #Color de texto
            cell.fill = PatternFill("solid", fgColor="4BACC6")
            cell.alignment = Alignment(horizontal="center")
            bd = Side(style='thin', color="000000")
            cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        
        for rows in ws['A2':max(ws.calculate_dimension()) + str(ws.max_row)]:
          for cell in rows:
            cell.font = Font(size=11)  #Color de texto
            bd = Side(style='thin', color="000000")
            cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        
        for rows in ws['A2': 'C' + str(ws.max_row)]:
          for cell in rows:
            cell.alignment = Alignment(horizontal="center")
            
        ws.auto_filter.ref='A1:' + max(ws.calculate_dimension()) + str(ws.max_row)
        wb.save('C:/Users/56975/Documents/proyecto_vicky.xlsx')
        return redirect('index')

    #except FileNotFoundError:
    #except:
    except Exception:
      wb = Workbook() # Crea un archivo necesario para trabajar.
      ws = wb.active # Crea libro de trabajo con una hoja de trabajo.

      # a.capitalize()  a.title()
      ws.title = (myDate.strftime("%B") + myDate.strftime("%Y")).upper() #Cambia el nombre de la hoja de trabajo.
      columnas = ['N°', 'LOCALIDAD', 'FECHA', 'NOMBRE', 'RUT', 'TIPO ATENCION', 'NACIONALIDAD', 'ETNIA', 'SEXO', 'RANGO ETARIO', 'EDAD']
      ws.append(columnas)
      datos = ['1', localidad, fecha, nombre, rut, tipo_atencion, nacionalidad, etnia, sexo, rango_etario, edad]
      ws.append(datos)

      for rows in ws['A1':'K1']:
        for cell in rows:
          cell.font = Font(color="000000", size=16, bold=True)  #Color de texto
          cell.fill = PatternFill("solid", fgColor="4BACC6")
          cell.alignment = Alignment(horizontal="center")
          bd = Side(style='thin', color="000000")
          cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)

      for rows in ws['A2':max(ws.calculate_dimension()) + str(ws.max_row)]:
        for cell in rows:
          cell.font = Font(size=11)  #Color de texto
          bd = Side(style='thin', color="000000")
          cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)

      for rows in ws['A2': 'C' + str(ws.max_row)]:
        for cell in rows:
          cell.alignment = Alignment(horizontal="center")

      #Filtro
      ws.auto_filter.ref='A1:' + max(ws.calculate_dimension()) + str(ws.max_row)
      wb.save('C:/Users/56975/Documents/proyecto_vicky.xlsx')
      return redirect('index')
  return render(request, "index.html", {'form': form})

def login_view(request):
  if request.user.is_authenticated:
    return redirect('index')
  #print(request.method)
  if request.method == 'POST':
    username = request.POST['username']
    password = request.POST['password']
    #username1 = request.POST.get('username')
    #password1 = request.POST.get('password')

    user = authenticate(username=username, password=password) # Retorna None en caso de no validar
    if user:
      login(request, user)
      print("Usuario autenticado")
      messages.success(request, 'Bienvenido {}'.format(user.username))
      return redirect('index')
    else:
      print("Usuario no autenticado")
      messages.error(request, 'Usuario o contraseña no validos')
  print("Adios!!")
  return render(request, 'users/login.html')

def logout_view(request):
  logout(request)
  messages.success(request, 'Sesión cerrada exitosamente')
  return redirect('login')

def register(request):
  if request.user.is_authenticated:
    return redirect('index')

  form = RegisterForm(request.POST or None)

  if request.method == 'POST' and form.is_valid():

    user = form.save()
    if user:
      login(request, user)
      messages.success(request, 'Usuario creado exitosamente')
      return redirect('index')

  return render(request, 'users/register.html', {'form': form})
